"""Append-only Excel feedback storage."""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from threading import Lock
from typing import Any

from openpyxl import Workbook, load_workbook


class FeedbackStorage:
    HEADERS = ["timestamp", "language", "tour_type", "rating_tour",
               "rating_explanation", "rating_robot", "comment"]

    def __init__(self, path: str | Path = "data/feedback/feedback.xlsx") -> None:
        self.path = Path(path)
        self._lock = Lock()

    def save(self, feedback: dict[str, Any]) -> Path:
        with self._lock:
            self.path.parent.mkdir(parents=True, exist_ok=True)
            if self.path.exists():
                workbook = load_workbook(self.path)
                sheet = workbook.active
            else:
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "Feedback"
                sheet.append(self.HEADERS)
                sheet.freeze_panes = "A2"
            sheet.append([
                datetime.now().astimezone().isoformat(timespec="seconds"),
                feedback.get("language"), feedback.get("tour_type"),
                feedback.get("rating_tour"), feedback.get("rating_explanation"),
                feedback.get("rating_robot"), feedback.get("comment", ""),
            ])
            for column in "ABCDEFG":
                sheet.column_dimensions[column].width = 20 if column != "G" else 42
            workbook.save(self.path)
        return self.path
